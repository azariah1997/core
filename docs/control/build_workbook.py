#!/usr/bin/env python3
"""Builds docs/control/data.xlsx - the Core Platform control-room workbook.

Run from anywhere: python3 docs/control/build_workbook.py
Edit the row data below, re-run, then reload docs/control/index.html.
"""
import base64
import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "data.xlsx")
HTML = os.path.join(HERE, "index.html")

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
DONE_FILL = PatternFill("solid", fgColor="D1FAE5")
PENDING_FILL = PatternFill("solid", fgColor="FEF3C7")
NEXT_FILL = PatternFill("solid", fgColor="DBEAFE")
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_C = Alignment(wrap_text=True, vertical="top", horizontal="center")

wb = openpyxl.Workbook()
wb.remove(wb.active)


def add_sheet(name, headers, rows, widths, freeze="A2", row_height=None):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_C
        cell.border = BORDER
    for r, row in enumerate(rows, start=2):
        ws.append(row)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = WRAP
        if row_height:
            ws.row_dimensions[r].height = row_height
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = freeze
    last_col = get_column_letter(len(headers))
    last_row = len(rows) + 1
    tbl = Table(displayName=name.replace(" ", "").replace("&", "").replace("-", "")[:30], ref=f"A1:{last_col}{last_row}")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tbl)
    return ws


# ---------------------------------------------------------------------------
# 1. OVERVIEW
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Overview")
ws.sheet_view.showGridLines = False
ws["A1"] = "Core Platform - Control Room"
ws["A1"].font = Font(size=20, bold=True, color="1F2937")
ws["A2"] = "A generic, multi-product backend platform (Go services + Next.js admin + Flutter mobile + Terraform/Helm/Docker infra), built phase-by-phase against a 30-phase roadmap."
ws["A2"].font = Font(size=11, italic=True, color="4B5563")
ws["A2"].alignment = Alignment(wrap_text=True)
ws.merge_cells("A2:F2")
ws.row_dimensions[2].height = 32

stats = [
    ("Phases complete", "28 / 30", "93%"),
    ("Backend services", "3", "core-api, realtime-gateway, worker"),
    ("Frontend apps", "1", "apps/admin (Next.js 15 App Router)"),
    ("Developer portal", "1", "platform/backstage (real @backstage/create-app, 27-entity catalog)"),
    ("Client SDKs", "3", "Go, TypeScript, Dart - packages/go/coresdk, packages/typescript/core-sdk, packages/flutter/core_sdk"),
    ("Shared Go packages", "2", "packages/go/platformkit, packages/go/coresdk"),
    ("Domain modules (core-api)", "24", "one per completed backend phase"),
    ("Live HTTP endpoints", "147", "see 'API Endpoints' sheet"),
    ("DB migrations", "22", "data/migrations/0001-0022"),
    ("Real infra dependencies", "10", "Postgres, Valkey, Keycloak, OpenFGA, MinIO, OpenSearch, Temporal, Kafka/Redpanda, Ollama, OTel/Prometheus/Grafana/Loki/Tempo"),
    ("Commits so far", "40", "see git log"),
]
r = 4
ws.cell(row=r, column=1, value="At a glance").font = Font(size=13, bold=True)
r += 1
for label, val, note in stats:
    ws.cell(row=r, column=1, value=label).font = Font(bold=True)
    ws.cell(row=r, column=2, value=val).font = Font(size=12, color="047857", bold=True)
    ws.cell(row=r, column=3, value=note).font = Font(italic=True, color="6B7280")
    r += 1

r += 1
ws.cell(row=r, column=1, value="How to read this workbook").font = Font(size=13, bold=True)
r += 1
notes = [
    "Roadmap - all 30 phases, one row each, status + what it delivers.",
    "Todo Checklist - granular task list (completed sub-items and still-to-do sub-items) so progress is trackable at a finer grain than 'phase done or not'.",
    "Modules - every backend/core-api/internal module: what it owns, its storage, its access rule.",
    "API Endpoints - every live HTTP route across all three services, with auth level.",
    "How To Test - copy-paste steps to bring the stack up locally and exercise it for real (mint a token, call endpoints, open a WebSocket).",
    "Infra & Credentials - every local service, its port, and its local-only credentials.",
]
for n in notes:
    ws.cell(row=r, column=1, value="• " + n).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1

ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 70

# ---------------------------------------------------------------------------
# 2. ROADMAP
# ---------------------------------------------------------------------------
roadmap_rows = [
    (1, "Done", "Platform Foundation", "config/logging/errors/health/correlation IDs/OTel/graceful shutdown - the shared platformkit every later phase builds on.", "9b9a65a"),
    (2, "Done", "Application Registry", "First real domain module: Application CRUD, cursor pagination, application.created/updated events.", "65a937d"),
    (3, "Done", "Identity", "Provider abstraction over Keycloak (JWT validation + Admin API), platform-owned IdentityID/UserID, identities linkage table.", "987f0cf"),
    (4, "Done", "Users", "Platform person/account, separate from Identity. Soft-delete lifecycle, auto-provision on first login.", "3f0289e"),
    (5, "Done", "Devices / Sessions", "Per-user device tracking, push-token storage, revoke/reactivate, raw tokens never exposed in responses.", "0762cfa"),
    (6, "Done", "Authorization / Access", "authz.Service.Can(subject, action, resource) - RBAC (Postgres) + fine-grained ReBAC (OpenFGA), the one entry point every module uses.", "892db7b"),
    (7, "Done", "Tenants / Organizations", "Tenant + Membership (owner/admin/member), scoped to an Application. Tenant-role checks answered from Membership data, not OpenFGA.", "9040905"),
    (8, "Done", "Relationships / Social Graph", "Generic RelationshipType/Relationship/Status graph: request/accept/decline/remove/block, no fixed 'friend/follower' semantics.", "542604c"),
    (9, "Done", "Groups / Circles", "Generic Group/GroupMember primitive (friend circles, teams, guilds - same shape), free-form product-defined member roles.", "02b5767"),
    (10, "Done", "Realtime Platform", "New realtime-gateway service: hand-rolled WebSocket (RFC6455), Redis-fanned pub/sub for cross-replica delivery, TTL-based presence.", "1eda4b3"),
    (11, "Done", "Messaging", "Conversation/Message/Delivery/ReadState/Reaction/Attachment, realtime delivery via a new shared rtbus package, durable storage in Postgres.", "a199c4c"),
    (12, "Done", "Notification Platform", "Notification/Template/Preference/Delivery across push/email/sms/in_app/realtime; only push is genuinely provider-tested locally (real device-token check).", "95b8e05"),
    (13, "Done", "File / Media Platform", "Presigned upload/download against real MinIO/S3, real MD5 checksum verification, ownership/permissions/retention/purge.", "9a39606"),
    (14, "Done", "Search Platform", "SearchDocument/SearchProvider over OpenSearch; event-driven indexing via a new worker service polling the outbox.", "6a4eafd"),
    (15, "Done", "Background Jobs", "Job/JobStatus/JobAttempt: immediate/scheduled/delayed/recurring/retry/dead-letter, executed by worker, never inline in HTTP handlers.", "d9221a4"),
    (16, "Done", "Workflows", "Temporal abstraction (Core Workflow API/SDK) for long-running, multi-step, compensating, scheduled processes - Temporal never exposed directly.", "59dfe77"),
    (17, "Done", "Feature Flags", "Feature/FeatureRule/FeatureEvaluation targeting app/env/user/tenant/platform/country/%/version, deterministic FNV-1a percentage bucketing.", "86925a5"),
    (18, "Done", "Remote Configuration", "Typed key/value store scoped by (AppID, Environment, Key), full change-audit trail, config.updated/deleted events.", "9a0ad7b"),
    (19, "Done", "Audit", "Central Audit Service: actor/action/resource/timestamp/correlation/app/tenant/device/metadata. Immutable by omission AND a DB trigger.", "2629e91"),
    (20, "Done", "Privacy", "Consent (append-only), preferences, retention policy, and cross-module ExportUserData/DeleteUserData via an Exporter/Deleter registry (users/devices/files/audit-export-only). Coordinated by a Temporal workflow core-api runs its own embedded worker for, since worker (separate module) can't reach core-api's internal packages.", "44299b2"),
    (21, "Done", "Trust & Safety", "Mute/Report->ModerationCase (deduplicated)/Suspension/Ban/Appeal - Block reuses Phase 8's relationships, not duplicated. A new requireActive middleware makes Suspension/Ban actually restrict access platform-wide from one place. New shared ratelimit package (Valkey-backed) gates report spam; a critical AbuseSignal auto-opens a case.", "7af4c58"),
    (22, "Done", "Billing / Entitlements", "Entitlement (platform truth, HasEntitlement) kept separate from Payment (provider transaction record). PaymentProvider abstraction + registry; billing/stripe implements Stripe's real webhook HMAC signature scheme, live-validated without a Stripe account via self-signed test payloads. Apple IAP/Google Play deliberately deferred - no sandbox credentials available.", "8fcc87f"),
    (23, "Done", "Analytics", "Generic event envelope, matching the roadmap's own field names verbatim. analytics_events is a landing buffer, never queried for analytics - a new worker pipeline batches unflushed rows as NDJSON into MinIO/S3 (a real data-lake landing pattern). Track is this platform's one open, unauthenticated write endpoint, IP-rate-limited instead of Bearer-gated.", "aa207ae"),
    (24, "Done", "AI Gateway", "Provider-neutral interface; real local inference via a new Ollama container (auto-pulled model, no vendor API key) - the one live-testable provider. Model routing, quotas (ratelimit reused again), real token/cost tracking, a genuine audit.Record per call, prompt/version metadata, timeouts, fallback. OpenAI/Anthropic/Google structurally supported, not implemented (no real credentials).", "d7324b9"),
    (25, "Done", "Admin Portal", "The control-plane UI (apps/admin, Next.js 15 App Router), always through service APIs, never direct DB queries - real Keycloak-authenticated session, live data on Users/Roles/Applications/Audit/Moderation/Feature Flags/Jobs/Configuration/System Health/Billing/AI Gateway, honest ComingSoon reasoning for every area whose backend endpoint is still self-scoped-only. Added authz's first HTTP surface and GET /v1/users (admin-wide listing) to close real gaps. Live-validated end to end with a real headless browser: login, dashboard, a full role grant/revoke round trip, logout.", "97091a7"),
    (26, "Done", "Backstage", "A real @backstage/create-app instance (platform/backstage) whose catalog (catalog/system.yaml) is 24 real entities - a Group, the System, Components for all 3 services + admin + 16 domain modules with real dependsOn/providesApis, and 2 API entities embedding the full real OpenAPI/AsyncAPI specs. Filled in 11 stub module READMEs, fixed 2 real YAML bugs in the OpenAPI spec, generated ~110 missing OpenAPI paths from the actual Go route registrations, and live-validated the whole catalog loads with zero processing errors via the real catalog API.", "b1bf7b9"),
    (27, "Done", "SDKs", "Go (packages/go/coresdk), TypeScript (packages/typescript/core-sdk), and Dart (packages/flutter/core_sdk) - auth/token refresh, typed API calls, errors, pagination, GET-only retries, a real realtime WebSocket client, device registration, in all three. Each proven by a real consumer: apps/admin migrated its entire API layer to the TS SDK; apps/mobile's login/profile flow uses the Dart SDK (and closed a real Phase 9 gap - make flutter-test now passes). 39 unit tests total, all against real local HTTP servers, plus live validation against real Keycloak/core-api/realtime-gateway for all three.", "12b9046"),
    (28, "Done", "Observability Completion", "New packages/go/platformkit/metrics gives every service a real /metrics endpoint and all 8 named dashboards real backing data (Kafka lag excluded - no real producer/consumer exists yet, an honest placeholder panel instead). logging.NewWithLoki ships structured logs directly to Loki's push API with real trace_id correlation. Grafana provisioning (previously mounted nowhere) now auto-loads Prometheus/Tempo/Loki datasources and a real dashboard; alerts.yml adds 6 live-evaluated Prometheus alerting rules. A real regression (broken WebSocket Hijack) was caught by the metrics middleware's own tests before ever reaching a running service.", "697ea4c"),
    (29, "Next up", "Platform Control Plane", "One view of applications/services/versions/environments/dependencies/deployments/DB ownership/events/API contracts/health/alerts/recent changes - every module discoverable from one place.", "-"),
    (30, "Pending", "AI Development Context", "Machine-readable context for AI agents - every module carries README, ownership metadata, dependencies, API contract, event contracts, DB ownership, so an agent (like this one) can safely work on any part of the platform.", "-"),
]
ws2 = add_sheet(
    "Roadmap",
    ["Phase", "Status", "Name", "What it delivers", "Commit"],
    roadmap_rows,
    widths=[7, 10, 26, 95, 10],
    row_height=46,
)
for r in range(2, len(roadmap_rows) + 2):
    status = ws2.cell(row=r, column=2).value
    fill = DONE_FILL if status == "Done" else (NEXT_FILL if status == "Next up" else PENDING_FILL)
    ws2.cell(row=r, column=2).fill = fill
    ws2.cell(row=r, column=2).alignment = WRAP_C
    ws2.cell(row=r, column=1).alignment = WRAP_C

# ---------------------------------------------------------------------------
# 3. TODO CHECKLIST (granular)
# ---------------------------------------------------------------------------
todo_rows = []


def done(phase, task, note=""):
    todo_rows.append((phase, "Done", task, note))


def pending(phase, task, note=""):
    todo_rows.append((phase, "Pending", task, note))


for n, name in [
    (1, "Platform Foundation"), (2, "Application Registry"), (3, "Identity"), (4, "Users"),
    (5, "Devices / Sessions"), (6, "Authorization / Access"), (7, "Tenants / Organizations"),
    (8, "Relationships / Social Graph"), (9, "Groups / Circles"), (10, "Realtime Platform"),
    (11, "Messaging"), (12, "Notification Platform"), (13, "File / Media Platform"),
    (14, "Search Platform"), (15, "Background Jobs"), (16, "Workflows"), (17, "Feature Flags"),
    (18, "Remote Configuration"), (19, "Audit"), (20, "Privacy"), (21, "Trust & Safety"),
    (22, "Billing / Entitlements"), (23, "Analytics"), (24, "AI Gateway"), (25, "Admin Portal"),
    (26, "Backstage"), (27, "SDKs"), (28, "Observability Completion"),
]:
    done(n, f"Phase {n}: {name} implemented, unit-tested, live-validated, documented, committed")

# Extra granular items worth tracking separately (known, documented gaps / deferred sub-scopes)
done(1, "Repo health baseline (make doctor / validate-all / go test / go build) run and passed")
done(2, "Environment / ApplicationVersion / ApplicationConfiguration", "Deferred - no consumer needs them yet, documented as intentional in applications/README.md")
done(5, "Separate Session table", "Deferred - Device.SessionStatus covers today's single-session-per-device need, documented in devices/README.md")
pending(9, "UserPreferences (named in Phase 4, no fields ever specified)", "Deferred to Phase 12+; notification preferences shipped, a generic preferences bag did not")
pending(9, "apps/mobile test/ directory (make flutter-test currently fails)", "Known gap since Phase 9's audit, not yet picked up")
pending(9, "apps/admin next@15 / sharp transitive advisories", "Fix is a Next.js major-version bump - separate, reviewable change")
done(20, "Privacy: consent (append-only history)")
done(20, "Privacy: user data export (ExportUserData registry - users/devices/files/audit)")
done(20, "Privacy: user deletion (DeleteUserData registry - users/devices/files; audit deliberately excluded)")
done(20, "Privacy: retention policy (declared, not scheduled - same documented gap as Phase 13's PurgeExpired)")
done(20, "Privacy: privacy preferences")
done(20, "Privacy: coordinate deletion through workflows/events - core-api's own embedded Temporal worker")
pending(20, "Privacy: notifications module as an Export/Delete participant", "Deliberately deferred to prove the registry makes this a 2-line addition later, not a framework change")
pending(20, "Privacy: relationships/groups/messaging as participants", "Deferred - shared multi-user data needs a redaction strategy this phase's per-user-in-isolation registry doesn't fit")
done(21, "Trust & Safety: Block", "Reused Phase 8's relationships (Status='blocked'), deliberately not duplicated")
done(21, "Trust & Safety: Mute")
done(21, "Trust & Safety: Report -> ModerationCase (deduplicated via partial unique index)")
done(21, "Trust & Safety: Suspension")
done(21, "Trust & Safety: Ban")
done(21, "Trust & Safety: Appeal (approving lifts the target restriction)")
done(21, "Trust & Safety: rate limiting (new platformkit/ratelimit, Valkey-backed)")
done(21, "Trust & Safety: spam protection (5 reports/hour per reporter)")
done(21, "Trust & Safety: abuse signals (critical severity auto-opens a case)")
done(21, "Trust & Safety: pluggable product-specific report reasons (free-form Reason field)")
done(21, "Trust & Safety: Suspension/Ban actually restrict access (new requireActive middleware, one entry point)")
pending(21, "Trust & Safety: WebSocket disconnect on suspension/ban", "Deferred - only HTTP access is gated; an existing realtime-gateway connection isn't forcibly closed")
pending(21, "Trust & Safety: aggregating lower-severity signals over a time window", "Deferred - a real abuse-detection engine is out of this phase's scope; only single-critical-signal escalation is implemented")
done(22, "Billing: Payment model (provider-specific transaction record)")
done(22, "Billing: Entitlement model (platform-facing truth, kept separate from Payment)")
done(22, "Billing: provider abstraction (PaymentProvider interface + registry, implemented first)")
done(22, "Billing: Stripe adapter", "Real HMAC-SHA256 webhook signature verification, live-validated without a Stripe account via self-signed test payloads")
pending(22, "Billing: Apple IAP adapter", "Deferred - needs real App Store Server API credentials this environment doesn't have")
pending(22, "Billing: Google Play adapter", "Deferred - needs real Play Developer API credentials this environment doesn't have")
done(23, "Analytics: generic event envelope (event_name/user_id/anonymous_id/app_id/session_id/timestamp/properties/context)")
done(23, "Analytics: keep operational DBs out of the analytics path (analytics_events is a landing buffer, never queried for analytics)")
done(23, "Analytics: pipeline foundation for ClickHouse/warehouse/data lake (worker batches NDJSON into MinIO/S3)")
pending(23, "Analytics: an actual ClickHouse/warehouse/data-lake query layer", "Out of scope - this phase builds the landing pipeline only; no real destination system exists in this environment to integrate against honestly")
done(24, "AI Gateway: provider-neutral interface (Provider: Name + Complete)")
done(24, "AI Gateway: local models / Ollama adapter", "Real local inference via a new Ollama container, auto-pulled model, no vendor API key - the one live-testable provider")
pending(24, "AI Gateway: OpenAI adapter", "Deferred - no real API key in this environment to validate against honestly")
pending(24, "AI Gateway: Anthropic adapter", "Deferred - no real API key in this environment to validate against honestly")
pending(24, "AI Gateway: Google adapter", "Deferred - no real API key in this environment to validate against honestly")
done(24, "AI Gateway: model routing (product-facing ModelAlias, never a vendor model name)")
done(24, "AI Gateway: quotas (ratelimit.Limiter reused a third time, 60/min/caller)")
done(24, "AI Gateway: token usage tracking")
done(24, "AI Gateway: cost tracking (real per-route price table; Ollama honestly priced at 0)")
done(24, "AI Gateway: audit integration (new AIGatewayAuditRecorder, a genuine audit.Record per call)")
done(24, "AI Gateway: prompt/version metadata (free-form PromptKey/PromptVersion)")
done(24, "AI Gateway: timeouts (context.WithTimeout per provider call)")
done(24, "AI Gateway: fallback (ordered provider/model chain, first success wins)")
pending(24, "AI Gateway: live token-quantity quota (as opposed to today's call-count quota)", "Deferred - Repository already has the data (Completion.TotalTokens); not wired as a live gate yet")
done(25, "Admin Portal: real Keycloak-authenticated session (Resource Owner Password Credentials grant), httpOnly session cookie")
done(25, "Admin Portal: visibility - Applications", "Real GET /v1/apps table")
done(25, "Admin Portal: visibility - Users (list + detail)", "New admin-wide GET /v1/users (cursor-paginated, platform.admin-gated)")
done(25, "Admin Portal: visibility - Roles/Permissions", "New authz HTTP surface (GET/POST /v1/authz/roles, POST /v1/authz/roles/revoke) with a full grant/revoke UI")
done(25, "Admin Portal: visibility - Feature Flags, Configuration, Jobs")
done(25, "Admin Portal: visibility - Moderation, Audit")
done(25, "Admin Portal: visibility - System Health", "Live GET /healthz against core-api/realtime-gateway/worker")
done(25, "Admin Portal: visibility - Billing, AI Gateway usage (bonus, not roadmap-named)")
pending(25, "Admin Portal: visibility - Sessions/Devices, Tenants, Relationships, Messages, Notifications, Files", "Deferred - each backing module only exposes a listMine-style self-scoped endpoint, not an admin-wide one; honest ComingSoon reasoning shown per area instead of fake rows")
pending(25, "Admin Portal: visibility - Deployments", "Deferred to Phase 29 - nothing in the platform tracks deployments as data yet")
done(26, "Backstage: integrate Backstage as developer portal", "Real @backstage/create-app instance at platform/backstage, catalog wired to real catalog/*.yaml + contracts/*/catalog-info.yaml, live-validated via the real catalog API (21 Components, 2 APIs, 1 System, 1 Group, 0 processing errors)")
done(26, "Backstage: per-module metadata - name/description/owner/lifecycle", "All 16 modules/*/module.yaml + README.md real (11 were still stub text before this phase)")
done(26, "Backstage: per-module metadata - repository/documentation", "Every Component links its module README; the System entity links VALIDATION.md and the new docs/RUNBOOKS.md")
done(26, "Backstage: per-module metadata - dependencies", "Component dependsOn mirrors modules/*/module.yaml's real dependency graph; Backstage computes reverse dependencyOf relations automatically, confirmed live")
done(26, "Backstage: per-module metadata - APIs", "2 API entities (OpenAPI, AsyncAPI), both with their full real spec embedded and live-loaded")
done(26, "Backstage: per-module metadata - events", "providesApis set only on the components whose code actually calls outbox.Record, cross-checked against the real Go source")
pending(26, "Backstage: per-module metadata - database ownership", "Deferred - no per-module DB-ownership annotation convention chosen yet; docs/control's Modules sheet tracks this today at a finer grain than catalog/system.yaml's 16 domains")
done(26, "Backstage: per-module metadata - dashboards", "core-platform.io/dashboard annotation on each service pointing at the real local Grafana instance - no per-service dashboard exists yet (that's Phase 28), so it points at the shared instance, documented as such")
done(26, "Backstage: per-module metadata - runbooks", "New docs/RUNBOOKS.md, linked from the System entity")
done(27, "SDKs: Go SDK", "packages/go/coresdk - 12 unit tests (httptest.Server), live-validated against real Keycloak/core-api/realtime-gateway")
done(27, "SDKs: TypeScript SDK", "packages/typescript/core-sdk - 13 unit tests (node:http), real consumer: apps/admin's entire lib/api.ts migrated to it")
done(27, "SDKs: Dart SDK", "packages/flutter/core_sdk (pure Dart, no Flutter dependency) - 13 unit tests (dart:io HttpServer), real consumer: apps/mobile's login/profile flow")
done(27, "SDKs: auth + token refresh in each SDK", "TokenSource (static + real Keycloak password-grant with skew-aware, concurrency-coalesced refresh) in all three")
done(27, "SDKs: pagination + safe retries", "Page/paginate/collectAll over the real {items,nextCursor} shape; GET-only retries by default on transient statuses, in all three")
done(27, "SDKs: realtime connection helper", "Reproduces realtime-gateway's real ws://.../ws?access_token=&deviceId= protocol and message shapes exactly, in all three - live-validated with real cross-connection fan-out")
done(27, "SDKs: correlation ID propagation", "X-Correlation-ID on every call, matching platformkit/correlation's real header name, in all three")
done(27, "SDKs: device registration helper", "DevicesRegister/devicesRegister in all three - realtime dial requires a real registered device id")
done(27, "SDKs: apps/mobile test/ directory (make flutter-test currently fails)", "Closed as a side effect - real widget tests added, discovered and worked around Flutter's TestWidgetsFlutterBinding network-mocking constraint live. Was pending since Phase 9.")
pending(27, "SDKs: full API coverage (all 146 endpoints typed) in each SDK", "Deferred - each SDK covers a representative 'core identity' slice (platform/identity/users/devices/applications) plus what its real consumer needs; the untyped request()/Do() escape hatch reaches everything else. Broader coverage is real, additive work, not a framework limitation.")
done(28, "Observability: structured logs everywhere", "logging.NewWithLoki ships every record to Loki's real push API directly; on by default locally (LOKI_PUSH_URL)")
done(28, "Observability: metrics everywhere", "New packages/go/platformkit/metrics - real GET /metrics (Prometheus exposition) on core-api, realtime-gateway, and worker")
done(28, "Observability: distributed traces everywhere", "Already real since Phase 1 (otelx) - confirmed still working live this phase")
done(28, "Observability: health checks everywhere", "Already real since Phase 1 (/livez /readyz /healthz) - confirmed still working live this phase")
done(28, "Observability: dashboard - API latency", "http_request_duration_seconds, labeled by route pattern (not raw path) to keep cardinality bounded")
done(28, "Observability: dashboard - error rate", "http_requests_total{status=~\"5..\"} / http_requests_total")
done(28, "Observability: dashboard - requests/sec", "rate(http_requests_total[5m])")
done(28, "Observability: dashboard - DB connections", "db_pool_connections, real pgxpool.Stat() polled on a ticker (pg.ReportStats) on all 3 services")
pending(28, "Observability: dashboard - Kafka lag", "Deliberately not implemented - no real Kafka/Redpanda producer or consumer exists in this codebase yet (documented since Phase 14); an honest placeholder panel explains this instead of a fabricated metric")
done(28, "Observability: dashboard - WebSocket connections", "realtime_ws_connections, live-validated going 0 -> 1 -> 0 across a real connect/hold/disconnect cycle against the running service")
done(28, "Observability: dashboard - Redis latency", "redis_command_duration_seconds via a real go-redis v9 Hook, InstrumentRedis")
done(28, "Observability: dashboard - notification failures", "notification_delivery_failures_total, incremented at notifications.Service.dispatch's two real failure paths")
done(28, "Observability: dashboard - background job failures", "job_failures_total, incremented only on jobrunner's real dead_letter transition (not every retryable attempt)")
done(28, "Observability: alerting rules", "6 real, live-evaluated Prometheus rules (infra/observability/alerts.yml) - ServiceDown live-confirmed firing then clearing on the real service lifecycle")
pending(28, "Observability: Alertmanager routing (Slack/PagerDuty/email)", "Deferred - no real notification-channel credentials in this environment, the same reasoning Stripe/AI vendor adapters have documented; Prometheus's own rule evaluation and API are real regardless")
pending(28, "Observability: per-service (not shared) Grafana dashboards", "Deferred - one shared 'Core Platform Overview' dashboard covers all 3 services today, not 9 separate roadmap-named dashboards")

for n, name, items in [
    (29, "Platform Control Plane", ["one view: applications/services/versions", "one view: environments/dependencies/deployments", "one view: DB ownership/events/API contracts", "one view: health/alerts/recent changes"]),
    (30, "AI Development Context", ["every module: README", "every module: ownership metadata", "every module: dependencies", "every module: API contract", "every module: event contracts", "every module: DB ownership"]),
]:
    for item in items:
        pending(n, f"{name}: {item}")

ws3 = add_sheet(
    "Todo Checklist",
    ["Phase", "Status", "Task", "Note"],
    todo_rows,
    widths=[7, 10, 70, 60],
    row_height=None,
)
for r in range(2, len(todo_rows) + 2):
    status = ws3.cell(row=r, column=2).value
    ws3.cell(row=r, column=2).fill = DONE_FILL if status == "Done" else PENDING_FILL
    ws3.cell(row=r, column=2).alignment = WRAP_C
    ws3.cell(row=r, column=1).alignment = WRAP_C

# ---------------------------------------------------------------------------
# 4. MODULES
# ---------------------------------------------------------------------------
module_rows = [
    ("applications", "core-api", "Application CRUD, the platform's own top-level tenant-of-tenants concept", "applications", "Open (read) / unauthenticated create in this phase's scope", "Phase 2"),
    ("identity", "core-api", "Auth provider abstraction (Keycloak today), token validation, identity<->user linkage", "identities", "Bearer token required for /me", "Phase 3"),
    ("users", "core-api", "Platform person/account, profile, soft-delete lifecycle", "users", "Self, or self-or-platform.admin for cross-user GET", "Phase 4"),
    ("devices", "core-api", "Per-user device + push-token registry, revoke/reactivate", "devices", "Self only", "Phase 5"),
    ("authz", "core-api", "Can(subject, action, resource) - RBAC + OpenFGA ReBAC, the one authorization entry point", "user_roles + OpenFGA store", "Self-service (HasRole) / role changes are the caller's own privileged action", "Phase 6"),
    ("tenants", "core-api", "Tenant + Membership (owner/admin/member), scoped to an Application", "tenants, tenant_memberships", "Member-only reads, owner/admin-only writes", "Phase 7"),
    ("relationships", "core-api", "Generic relationship graph: request/accept/decline/remove/block", "relationships", "Participant-only", "Phase 8"),
    ("groups", "core-api", "Generic Group/GroupMember, free-form roles, IsManager structural bit", "groups, group_members", "Member-only reads, manager-only writes", "Phase 9"),
    ("realtime-gateway (service)", "realtime-gateway", "WebSocket connections, channels, presence, direct/broadcast delivery", "none (Redis/Valkey only)", "Bearer token required at handshake", "Phase 10"),
    ("messaging", "core-api", "Conversation/Message/Delivery/ReadState/Reaction/Attachment", "conversations, conversation_members, messages, message_*", "Conversation-member-only", "Phase 11"),
    ("notifications", "core-api", "Notification/Template/Preference/Delivery across 5 channels", "notifications, notification_requests, notification_preferences", "Self-or-platform.admin to send; templates are platform.admin-only", "Phase 12"),
    ("files", "core-api", "Presigned upload/download, ownership, retention, checksum", "files", "Owner-or-platform.admin, visibility-scoped downloads", "Phase 13"),
    ("search", "core-api", "SearchDocument/SearchProvider over OpenSearch", "none (OpenSearch only)", "Open query; admin-only manual index/delete", "Phase 14"),
    ("worker: indexer", "worker", "Event-driven OpenSearch indexer, polls outbox_events", "reads outbox_events", "n/a (background)", "Phase 14"),
    ("jobs", "core-api", "Job/JobStatus/JobAttempt - immediate/scheduled/delayed/recurring/retry/dead-letter", "jobs, job_attempts", "Owner-or-platform.admin", "Phase 15"),
    ("worker: jobrunner", "worker", "Claims and executes due jobs (Echo, Webhook handlers)", "reads/writes jobs, job_attempts", "n/a (background)", "Phase 15"),
    ("workflows", "core-api", "Core Workflow API/SDK over Temporal (Start/Describe/Signal only)", "workflow_runs (ownership only - state lives in Temporal)", "Owner-or-platform.admin", "Phase 16"),
    ("worker: workflows", "worker", "ApprovalWorkflow, PingWebhookWorkflow (Temporal worker)", "none (Temporal-native state)", "n/a (background)", "Phase 16"),
    ("features", "core-api", "Feature/FeatureRule/FeatureEvaluation, deterministic percentage rollout", "features, feature_rules", "Open evaluate; platform.admin-only manage", "Phase 17"),
    ("remoteconfig", "core-api", "Typed KV config store, full audit trail per change", "config_entries, config_changes", "Open read; platform.admin-only write/history", "Phase 18"),
    ("audit", "core-api", "Central, immutable audit trail (API-omission + DB trigger enforced)", "audit_events", "Open to record; platform.admin-only to read", "Phase 19"),
    ("privacy", "core-api", "Consent, preferences, retention policy, cross-module export/delete via a registry", "privacy_consents, privacy_preferences, retention_policies, data_export_requests, data_deletion_requests", "Self for own data; platform.admin for retention policies", "Phase 20"),
    ("privacy: embedded Temporal worker", "core-api", "Runs privacy's own export/deletion workflows in-process (task queue privacy-tasks)", "none (Temporal-native state)", "n/a (background, inside core-api's own process)", "Phase 20"),
    ("trustsafety", "core-api", "Mute/Report/ModerationCase/Suspension/Ban/Appeal/AbuseSignal - Block reuses Phase 8's relationships", "mutes, moderation_cases, reports, suspensions, bans, appeals, abuse_signals", "Self for own mutes/appeals; moderator-or-admin for case/suspension/ban management", "Phase 21"),
    ("billing", "core-api", "Entitlement (platform truth) kept separate from Payment (provider transaction record); PaymentProvider abstraction + registry", "entitlements, payments", "Self for own entitlements/payments; platform.admin for manual grants", "Phase 22"),
    ("billing/stripe", "core-api", "Real Stripe webhook HMAC-SHA256 signature verification - the one live-testable PaymentProvider this phase ships", "none (stateless verification)", "n/a (provider-signature-authenticated, not a Bearer route)", "Phase 22"),
    ("analytics", "core-api", "Generic event envelope ingestion; analytics_events is a landing buffer, never a query surface", "analytics_events", "Open (IP-rate-limited) to track; platform.admin-only debug listing", "Phase 23"),
    ("worker: analyticspipeline", "worker", "Claims unflushed events, batches as NDJSON, writes to MinIO/S3 - the ClickHouse/warehouse/data-lake landing pattern", "reads/writes analytics_events; writes MinIO/S3", "n/a (background)", "Phase 23"),
    ("aigateway", "core-api", "Provider-neutral completions API - model routing, quotas, cost tracking, audit, timeouts, fallback", "ai_completions", "Self for own usage; open (quota-limited) to complete", "Phase 24"),
    ("aigateway/ollama", "core-api", "Real local-inference Provider adapter - the one live-testable Provider this phase ships", "none (stateless HTTP client to a real Ollama container)", "n/a (internal provider adapter, not a route)", "Phase 24"),
]
ws4 = add_sheet(
    "Modules",
    ["Module", "Service", "Responsibility", "Storage", "Access model", "Introduced"],
    module_rows,
    widths=[22, 16, 58, 32, 42, 11],
    row_height=32,
)

# ---------------------------------------------------------------------------
# 5. API ENDPOINTS
# ---------------------------------------------------------------------------
def E(method, path, module, auth, note=""):
    return (method, path, module, auth, note)


endpoint_rows = [
    E("GET", "/livez", "platform", "Open", "process alive"),
    E("GET", "/readyz", "platform", "Open", "dependencies ready"),
    E("GET", "/healthz", "platform", "Open", "aggregated health"),
    E("GET", "/metrics", "platform", "Open", "Prometheus exposition (Phase 28) - real on core-api, realtime-gateway, and worker"),
    E("GET", "/v1/platform", "platform", "Open", "platform name/env/version"),
    E("POST", "/v1/apps", "applications", "Open (this phase's scope)", ""),
    E("GET", "/v1/apps", "applications", "Open", "cursor-paginated list"),
    E("GET", "/v1/apps/{id}", "applications", "Open", ""),
    E("PATCH", "/v1/apps/{id}", "applications", "Open (this phase's scope)", ""),
    E("GET", "/v1/identity/me", "identity", "Authenticated", "resolves the caller's Identity"),
    E("GET", "/v1/users/me", "users", "Authenticated (self)", ""),
    E("PATCH", "/v1/users/me", "users", "Authenticated (self)", ""),
    E("GET", "/v1/users/{id}", "users", "Self or platform.admin", ""),
    E("GET", "/v1/users", "users", "platform.admin", "Phase 25: cursor-paginated admin-wide listing, new to close a real Admin Portal gap"),
    E("GET", "/v1/authz/roles", "authz", "Self or platform.admin", "Phase 25: authz's first-ever HTTP surface"),
    E("POST", "/v1/authz/roles", "authz", "platform.admin", "assign a role to a user"),
    E("POST", "/v1/authz/roles/revoke", "authz", "platform.admin", "revoke a role from a user"),
    E("GET", "/v1/users/me/devices", "devices", "Authenticated (self)", ""),
    E("POST", "/v1/users/me/devices", "devices", "Authenticated (self)", "register/upsert"),
    E("DELETE", "/v1/users/me/devices/{id}", "devices", "Authenticated (self)", "soft-revoke"),
    E("GET", "/v1/tenants", "tenants", "Authenticated", "scoped to caller's own tenants"),
    E("POST", "/v1/tenants", "tenants", "Authenticated", "creator becomes owner"),
    E("GET", "/v1/tenants/{id}", "tenants", "Member", ""),
    E("PATCH", "/v1/tenants/{id}", "tenants", "Owner or admin", ""),
    E("GET", "/v1/tenants/{id}/members", "tenants", "Member", ""),
    E("POST", "/v1/tenants/{id}/members", "tenants", "Owner or admin", ""),
    E("DELETE", "/v1/tenants/{id}/members/{userId}", "tenants", "Owner/admin, or self (leave)", ""),
    E("GET", "/v1/relationships", "relationships", "Participant", ""),
    E("GET", "/v1/relationships/{id}", "relationships", "Participant", ""),
    E("POST", "/v1/relationships", "relationships", "Authenticated", "request"),
    E("POST", "/v1/relationships/block", "relationships", "Authenticated", ""),
    E("POST", "/v1/relationships/{id}/accept", "relationships", "Target only", ""),
    E("POST", "/v1/relationships/{id}/decline", "relationships", "Target only", ""),
    E("DELETE", "/v1/relationships/{id}", "relationships", "Participant", "remove/unfriend"),
    E("GET", "/v1/groups", "groups", "Member", ""),
    E("POST", "/v1/groups", "groups", "Authenticated", "creator becomes manager"),
    E("GET", "/v1/groups/{id}", "groups", "Member", ""),
    E("PATCH", "/v1/groups/{id}", "groups", "Manager", ""),
    E("GET", "/v1/groups/{id}/members", "groups", "Member", ""),
    E("POST", "/v1/groups/{id}/members", "groups", "Manager", ""),
    E("PATCH", "/v1/groups/{id}/members/{userId}", "groups", "Manager", "promote/demote IsManager"),
    E("DELETE", "/v1/groups/{id}/members/{userId}", "groups", "Manager, or self (leave)", ""),
    E("WS", "/ws (realtime-gateway)", "realtime-gateway", "Bearer token at handshake", "subscribe/publish/broadcast/direct message"),
    E("GET", "/v1/conversations", "messaging", "Member", ""),
    E("POST", "/v1/conversations", "messaging", "Authenticated", "direct convos dedup automatically"),
    E("GET", "/v1/conversations/{id}", "messaging", "Member", ""),
    E("GET", "/v1/conversations/{id}/members", "messaging", "Member", ""),
    E("POST", "/v1/conversations/{id}/members", "messaging", "Member (group only)", ""),
    E("DELETE", "/v1/conversations/{id}/members/{userId}", "messaging", "Self only (leave)", ""),
    E("GET", "/v1/conversations/{id}/messages", "messaging", "Member", "cursor pagination"),
    E("POST", "/v1/conversations/{id}/messages", "messaging", "Member", "pushes over realtime too"),
    E("GET", "/v1/conversations/{id}/messages/{messageId}", "messaging", "Member", ""),
    E("POST", "/v1/conversations/{id}/messages/{messageId}/delivered", "messaging", "Member", "idempotent ack"),
    E("GET", "/v1/conversations/{id}/messages/{messageId}/reactions", "messaging", "Member", ""),
    E("POST", "/v1/conversations/{id}/messages/{messageId}/reactions", "messaging", "Member", "idempotent add"),
    E("DELETE", "/v1/conversations/{id}/messages/{messageId}/reactions/{type}", "messaging", "Member", ""),
    E("GET", "/v1/conversations/{id}/read/{userId}", "messaging", "Member", ""),
    E("PUT", "/v1/conversations/{id}/read", "messaging", "Member", ""),
    E("GET", "/v1/notifications", "notifications", "Authenticated (self)", ""),
    E("POST", "/v1/notifications", "notifications", "Self or platform.admin", "send"),
    E("GET", "/v1/notifications/{id}", "notifications", "Self or platform.admin", ""),
    E("GET", "/v1/notifications/{id}/deliveries", "notifications", "Self or platform.admin", ""),
    E("POST", "/v1/notifications/{id}/deliveries/{deliveryId}/retry", "notifications", "Self or platform.admin", ""),
    E("GET", "/v1/notification-preferences", "notifications", "Authenticated (self)", ""),
    E("PUT", "/v1/notification-preferences", "notifications", "Authenticated (self)", "category opt-outs"),
    E("GET", "/v1/notification-preferences/quiet-hours", "notifications", "Authenticated (self)", ""),
    E("PUT", "/v1/notification-preferences/quiet-hours", "notifications", "Authenticated (self)", ""),
    E("POST", "/v1/notification-templates", "notifications", "platform.admin", ""),
    E("GET", "/v1/notification-templates/{appId}/{key}", "notifications", "platform.admin", ""),
    E("PATCH", "/v1/notification-templates/{appId}/{key}", "notifications", "platform.admin", ""),
    E("GET", "/v1/files", "files", "Owner or platform.admin", ""),
    E("POST", "/v1/files", "files", "Authenticated", "requests a presigned upload URL"),
    E("POST", "/v1/files/{id}/confirm", "files", "Owner", "verifies real size + MD5 checksum"),
    E("GET", "/v1/files/{id}", "files", "Owner, or public if visibility allows", ""),
    E("GET", "/v1/files/{id}/download", "files", "Owner, or public if visibility allows", "presigned GET"),
    E("DELETE", "/v1/files/{id}", "files", "Owner or platform.admin", "also deletes from storage"),
    E("POST", "/v1/files/purge-expired", "files", "platform.admin", ""),
    E("GET", "/v1/search", "search", "Open", ""),
    E("POST", "/v1/search/index", "search", "platform.admin", "manual index (normally event-driven)"),
    E("DELETE", "/v1/search/index/{type}/{id}", "search", "platform.admin", ""),
    E("GET", "/v1/jobs", "jobs", "Owner or platform.admin", ""),
    E("POST", "/v1/jobs", "jobs", "Authenticated", "enqueue only - never runs inline"),
    E("GET", "/v1/jobs/{id}", "jobs", "Owner or platform.admin", ""),
    E("GET", "/v1/jobs/{id}/attempts", "jobs", "Owner or platform.admin", ""),
    E("POST", "/v1/workflows", "workflows", "Authenticated", "starts a Temporal workflow"),
    E("GET", "/v1/workflows/{id}", "workflows", "Owner or platform.admin", "live-queried from Temporal"),
    E("POST", "/v1/workflows/{id}/signal", "workflows", "Owner or platform.admin", ""),
    E("GET", "/v1/features", "features", "platform.admin", ""),
    E("POST", "/v1/features", "features", "platform.admin", ""),
    E("GET", "/v1/features/{appId}/{key}", "features", "platform.admin", ""),
    E("PATCH", "/v1/features/{appId}/{key}", "features", "platform.admin", "kill switch etc"),
    E("POST", "/v1/features/{appId}/{key}/evaluate", "features", "Open", "isEnabled(feature, context)"),
    E("GET", "/v1/features/{appId}/{key}/rules", "features", "platform.admin", ""),
    E("POST", "/v1/features/{appId}/{key}/rules", "features", "platform.admin", ""),
    E("PATCH", "/v1/features/{appId}/{key}/rules/{ruleId}", "features", "platform.admin", ""),
    E("DELETE", "/v1/features/{appId}/{key}/rules/{ruleId}", "features", "platform.admin", ""),
    E("GET", "/v1/config", "remoteconfig", "Open", "list entries"),
    E("GET", "/v1/config/{appId}/{environment}/{key}", "remoteconfig", "Open", ""),
    E("PUT", "/v1/config/{appId}/{environment}/{key}", "remoteconfig", "platform.admin", "writes an audited Change row"),
    E("DELETE", "/v1/config/{appId}/{environment}/{key}", "remoteconfig", "platform.admin", ""),
    E("GET", "/v1/config/{appId}/{environment}/{key}/history", "remoteconfig", "platform.admin", ""),
    E("POST", "/v1/audit", "audit", "Authenticated (actor forced server-side)", ""),
    E("GET", "/v1/audit", "audit", "platform.admin", ""),
    E("GET", "/v1/audit/{id}", "audit", "platform.admin", ""),
    E("POST", "/v1/privacy/consent", "privacy", "Authenticated (self)", "append-only history"),
    E("GET", "/v1/privacy/consent", "privacy", "Authenticated (self)", ""),
    E("GET", "/v1/privacy/preferences", "privacy", "Authenticated (self)", ""),
    E("PUT", "/v1/privacy/preferences", "privacy", "Authenticated (self)", ""),
    E("POST", "/v1/privacy/retention-policies", "privacy", "platform.admin", ""),
    E("GET", "/v1/privacy/retention-policies", "privacy", "platform.admin", ""),
    E("POST", "/v1/privacy/export", "privacy", "Authenticated (self)", "starts a Temporal workflow"),
    E("GET", "/v1/privacy/export/{id}", "privacy", "Self or platform.admin", "status is read from Postgres, not Temporal"),
    E("GET", "/v1/privacy/export/{id}/download", "privacy", "Self or platform.admin", "presigned MinIO/S3 GET"),
    E("POST", "/v1/privacy/delete", "privacy", "Authenticated (self)", "starts a Temporal workflow"),
    E("GET", "/v1/privacy/delete/{id}", "privacy", "Self or platform.admin", "a self-deleted user must use an admin to check this"),
    E("POST", "/v1/trustsafety/mutes", "trustsafety", "Authenticated (self)", "one-directional, silent"),
    E("GET", "/v1/trustsafety/mutes", "trustsafety", "Authenticated (self)", "exempt from requireActive - always reachable"),
    E("DELETE", "/v1/trustsafety/mutes/{userId}", "trustsafety", "Authenticated (self)", ""),
    E("POST", "/v1/trustsafety/reports", "trustsafety", "Authenticated (self)", "rate-limited 5/hour; dedupes onto one case"),
    E("GET", "/v1/trustsafety/reports/{id}", "trustsafety", "Reporter or moderator/admin", ""),
    E("GET", "/v1/trustsafety/cases", "trustsafety", "Moderator or platform.admin", "?status="),
    E("GET", "/v1/trustsafety/cases/{id}", "trustsafety", "Moderator or platform.admin", ""),
    E("GET", "/v1/trustsafety/cases/{id}/reports", "trustsafety", "Moderator or platform.admin", ""),
    E("PATCH", "/v1/trustsafety/cases/{id}/assign", "trustsafety", "Moderator or platform.admin", ""),
    E("PATCH", "/v1/trustsafety/cases/{id}", "trustsafety", "Moderator or platform.admin", "resolve/dismiss"),
    E("POST", "/v1/trustsafety/suspensions", "trustsafety", "Moderator or platform.admin", "temporary, EndsAt required"),
    E("GET", "/v1/trustsafety/suspensions", "trustsafety", "Self or moderator/admin", "?userId="),
    E("POST", "/v1/trustsafety/suspensions/{id}/lift", "trustsafety", "Moderator or platform.admin", ""),
    E("POST", "/v1/trustsafety/bans", "trustsafety", "Moderator or platform.admin", "permanent until lifted"),
    E("GET", "/v1/trustsafety/bans", "trustsafety", "Self or moderator/admin", "?userId="),
    E("POST", "/v1/trustsafety/bans/{id}/lift", "trustsafety", "Moderator or platform.admin", ""),
    E("POST", "/v1/trustsafety/appeals", "trustsafety", "Authenticated (self, must own target)", "exempt from requireActive - always reachable"),
    E("GET", "/v1/trustsafety/appeals", "trustsafety", "Self or moderator/admin", "exempt from requireActive - always reachable"),
    E("POST", "/v1/trustsafety/appeals/{id}/review", "trustsafety", "Moderator or platform.admin", "approving lifts the target restriction"),
    E("POST", "/v1/trustsafety/signals", "trustsafety", "Authenticated", "critical severity auto-opens a case"),
    E("GET", "/v1/trustsafety/signals", "trustsafety", "Moderator or platform.admin", "?resourceType=&resourceId="),
    E("GET", "/v1/billing/entitlements", "billing", "Self or platform.admin", "?userId="),
    E("GET", "/v1/billing/entitlements/check/{key}", "billing", "Authenticated (self)", "the literal 'do I have entitlement X' check"),
    E("POST", "/v1/billing/entitlements", "billing", "platform.admin", "manual grant; source always \"manual:<adminId>\""),
    E("POST", "/v1/billing/entitlements/{id}/revoke", "billing", "platform.admin", ""),
    E("GET", "/v1/billing/payments", "billing", "Self or platform.admin", "?userId="),
    E("POST", "/v1/billing/webhooks/{provider}", "billing", "Provider webhook signature (no Bearer token)", "the one route on this platform with no requireUser/requireActive"),
    E("POST", "/v1/analytics/events", "analytics", "Open (IP-rate-limited, no Bearer token)", "the platform's one deliberately unauthenticated write endpoint"),
    E("GET", "/v1/analytics/events", "analytics", "platform.admin", "operational debug view only, not for analytical queries"),
    E("POST", "/v1/ai/completions", "aigateway", "Authenticated (quota-limited)", "real local inference via Ollama; never call a vendor model name directly"),
    E("GET", "/v1/ai/usage", "aigateway", "Self or platform.admin", "?userId=; tokens/cost/latency, never prompt or response text"),
    E("GET", "/v1/ai/models", "aigateway", "Authenticated", "lists registered model aliases, never vendor model names"),
]
ws5 = add_sheet(
    "API Endpoints",
    ["Method", "Path", "Module", "Auth", "Notes"],
    endpoint_rows,
    widths=[8, 46, 16, 32, 42],
)

# ---------------------------------------------------------------------------
# 6. HOW TO TEST
# ---------------------------------------------------------------------------
test_rows = [
    ("1", "Bring up local infra", "make local-up", "Starts Postgres, Valkey, Keycloak, OpenFGA, MinIO, OpenSearch, Temporal, Kafka/Redpanda + the observability stack (13 containers). Idempotent, one command."),
    ("2", "Check everything is healthy", "docker ps", "Every container should show 'Up' (Postgres shows '(healthy)'). Re-run make local-up if anything is missing - Keycloak/OpenFGA have no persistent volume, so realm/store data is re-imported on every restart."),
    ("3", "Build the services", "make build", "Compiles core-api, realtime-gateway, worker, and platformkit. Should exit 0 with no output."),
    ("4", "Run the test suite", "make test", "Runs every module's unit tests against in-memory fakes. Should be all 'ok'."),
    ("5", "Start core-api", "set -a && source .env && set +a && go run ./backend/core-api/cmd/server", "Listens on :8080 (see Infra sheet). Watch its stdout for structured JSON logs."),
    ("6", "Start realtime-gateway (optional)", "set -a && source .env && set +a && go run ./backend/realtime-gateway/cmd/server", "Listens on :8090. Only needed to test WebSocket/presence features."),
    ("7", "Start worker (optional)", "set -a && source .env && set +a && go run ./backend/worker/cmd/worker", "Runs the job runner, search indexer, and Temporal workflow worker. Only needed to test jobs/search/workflows end to end."),
    ("8", "Confirm the API is up", "curl -s http://localhost:8080/healthz", "Should return {\"status\":\"ok\",...} with every dependency 'ok': true."),
    ("9", "Mint a real login token", "curl -s -X POST http://localhost:8081/realms/core/protocol/openid-connect/token -d grant_type=password -d client_id=core-platform -d username=demo -d password=demo | python3 -c \"import sys,json;print(json.load(sys.stdin)['access_token'])\"", "The realm ships one seeded user, 'demo' / 'demo'. Save the output as $TOKEN - it's a real JWT good for 300 seconds."),
    ("10", "Call an authenticated endpoint", "curl -s http://localhost:8080/v1/users/me -H \"Authorization: Bearer $TOKEN\"", "First call auto-provisions a platform User linked to the Keycloak identity. Re-run it - you get back the same user."),
    ("11", "Create a second, non-admin test user", "Use Keycloak's Admin API (admin/admin against the master realm) to POST a user to /admin/realms/core/users - see the Audit phase's live validation for the exact curl sequence.", "Needed for anything that checks cross-user permissions (403-for-a-stranger tests)."),
    ("12", "Grant platform.admin to a user", "curl -s -X POST http://localhost:8080/v1/authz/roles -H \"Authorization: Bearer $TOKEN\" -H 'Content-Type: application/json' -d '{\"userId\":\"<id>\",\"role\":\"platform.admin\"}'", "Phase 25 gave authz a real HTTP surface - $TOKEN's own user must already be platform.admin (the seeded 'demo' user is, in this environment). Also doable from the Admin Portal's Users -> role management UI. The very first admin in a brand-new environment still needs a one-time throwaway Go program (AssignRole has no caller-privilege check at the service layer, only at HTTP) - the same bootstrap problem every real RBAC system has."),
    ("13", "Exercise a module end-to-end", "e.g. POST /v1/apps, then GET /v1/apps, then GET /v1/apps/{id} with the returned id", "Pick any row from the 'API Endpoints' sheet and try it - every route takes/returns plain JSON. Look at each module's own README.md (backend/core-api/internal/<module>/README.md) for the exact request/response shapes and the scoping decisions behind them."),
    ("14", "Watch a WebSocket live", "wscat -c \"ws://localhost:8090/ws?access_token=$TOKEN\" (or any RFC6455 client)", "You should receive a {\"type\":\"connected\",...} frame immediately. Subscribe to a channel, then publish from a second connection to see fan-out."),
    ("15", "Watch the database directly", "docker exec -it docker-postgres-1 psql -U core -d core", "Useful for sanity-checking: SELECT * FROM outbox_events ORDER BY id DESC LIMIT 20; or SELECT * FROM audit_events ORDER BY occurred_at DESC LIMIT 20;"),
    ("16", "Read the validation history", "VALIDATION.md (repo root)", "Every phase's actual live-validation pass is written up there in detail - exact curl sequences, real bugs found and fixed, and what was deliberately left out of scope."),
    ("17", "Run the smoke test", "make smoke", "A scripted end-to-end pass (scripts/smoke.sh) against the real running services - the fastest single command to confirm the whole stack still works after a change."),
    ("18", "Skip the curl entirely", "Open this dashboard's API Console tab", "Pick an endpoint, fetch a token via the built-in Keycloak password-grant form (or paste one you already have), and send the request - the real response comes back right in the page. core-api's CORS middleware (internal/api/cors.go) is what makes this possible; it must be running a build from Phase 21 or later."),
    ("19", "Run the actual Admin Portal", "cd apps/admin && npm install && npm run dev", "Then open http://localhost:3000 and sign in with demo/demo. Real UI over real data (Phase 25) - Users, Roles, Audit, Moderation, Feature Flags, Jobs, Configuration, System Health, Billing, AI Gateway. Needs core-api (and ideally realtime-gateway/worker) already running per steps 5-7."),
    ("20", "Browse the real developer-portal catalog", "cd platform/backstage && yarn install && yarn start", "Needs Node 20/22 and a real Yarn install - see platform/backstage/README.md if the versions on $PATH don't satisfy Backstage's own prerequisite check. Then open http://localhost:3000 (stop the Admin Portal first, they share a port) and browse the Catalog - every module/service/app is a real entity (Phase 26), not example fixture data."),
    ("21", "Test the Go SDK", "cd packages/go/coresdk && go test ./...", "12 unit tests against a real httptest.Server (Phase 27) - no mocking library."),
    ("22", "Test the TypeScript SDK", "cd packages/typescript/core-sdk && npm install && npm test", "13 unit tests against a real node:http server. make sdk-ts-build must run before apps/admin can resolve @core-platform/sdk - make admin-install/admin-build already depend on it."),
    ("23", "Test the Dart SDK", "cd packages/flutter/core_sdk && dart pub get && dart test", "13 unit tests against a real dart:io HttpServer. apps/mobile depends on this package via a path: dependency in its pubspec.yaml."),
    ("24", "Look at real dashboards, metrics, logs, traces, and alerts", "Open http://localhost:3000 (admin/admin) -> Core Platform folder -> Core Platform Overview", "Needs core-api/realtime-gateway/worker actually running (steps 5-7). Every panel is real data (Phase 28) - API latency, error rate, requests/sec, DB connections, WebSocket connections, Redis latency, notification/job failures. Prometheus alerts: http://localhost:9090/alerts. Loki/Tempo: query via Grafana's Explore tab."),
]
ws6 = add_sheet(
    "How To Test",
    ["Step", "What", "Command / Where", "Notes"],
    test_rows,
    widths=[6, 30, 70, 60],
    row_height=48,
)

# ---------------------------------------------------------------------------
# 7. INFRA & CREDENTIALS
# ---------------------------------------------------------------------------
infra_rows = [
    ("core-api", "Go service", "http://localhost:8080", "The main platform API - almost every module in this workbook lives here.", "n/a"),
    ("admin (apps/admin)", "Next.js app", "http://localhost:3000", "The Admin Portal (Phase 25) - real UI over core-api, session-based auth via Keycloak.", "sign in with demo/demo (same as core realm)"),
    ("backstage (platform/backstage)", "Backstage app", "http://localhost:3000 (frontend), :7007 (backend)", "The developer portal (Phase 26) - real catalog of every module/service/app. Shares apps/admin's default frontend port 3000, so only run one of the two at a time locally unless one's app-config.yaml is repointed at a different port.", "guest auth only, no credentials"),
    ("realtime-gateway", "Go service", "ws://localhost:8090", "WebSocket connections, presence, channel fan-out.", "n/a"),
    ("worker", "Go service", "http://localhost:8091 (health only)", "Background job runner, search indexer, Temporal workflow worker, analytics NDJSON pipeline.", "n/a"),
    ("Postgres", "Docker container", "localhost:5432", "System of record for every domain module.", "core / core, db 'core'"),
    ("Valkey (Redis-compatible)", "Docker container", "localhost:6379", "Realtime pub/sub fan-out, presence (TTL-based).", "no auth locally"),
    ("Keycloak", "Docker container", "http://localhost:8081", "Identity provider. Realm 'core' is re-imported on every restart (no persistent volume).", "admin/admin (master realm); demo/demo (core realm, seeded user)"),
    ("OpenFGA", "Docker container", "http://localhost:8082", "Fine-grained ReBAC engine behind authz.Service.Can. In-memory store, self-heals its store/model on every restart.", "no auth locally"),
    ("MinIO (S3-compatible)", "Docker container", "http://localhost:9000", "Object storage backing the files module.", "minio / minio123, bucket 'core-platform'"),
    ("OpenSearch", "Docker container", "http://localhost:9200", "Backs the search module, indexed by worker's indexer.", "no auth locally"),
    ("Temporal", "Docker container", "localhost:7233", "Durable workflow execution engine behind the workflows module.", "no auth locally"),
    ("Kafka / Redpanda", "Docker container", "localhost:9092", "Health-checked today; no producer wired yet (outbox-to-Kafka relay is a documented future gap).", "no auth locally"),
    ("Ollama", "Docker container", "http://localhost:11434", "Real local LLM inference behind the AI Gateway - model qwen2.5:0.5b auto-pulled on every start, no vendor API key needed.", "no auth locally"),
    ("Prometheus", "Docker container", "http://localhost:9090", "Scrapes a real GET /metrics on core-api/realtime-gateway/worker (Phase 28) and evaluates alerts.yml's 6 real alerting rules - api/v1/targets and api/v1/rules are both genuinely populated, not just running empty.", "no auth locally"),
    ("Grafana", "Docker container", "http://localhost:3000", "Provisioned (Phase 28) - Prometheus/Tempo/Loki datasources and the real 'Core Platform Overview' dashboard auto-load on container start, previously mounted nowhere.", "admin/admin"),
    ("Loki", "Docker container", "http://localhost:3100", "Receives structured logs pushed directly by logging.NewWithLoki (Phase 28) - no log-shipping agent, since these services run as bare host processes Promtail-style container-log-tailing would never see.", "no auth locally"),
    ("Tempo", "Docker container", "http://localhost:3200", "Receives real traces via otel-collector (Phase 1).", "no auth locally"),
    ("OTel Collector", "Docker container", "http://localhost:4318 (OTLP/HTTP)", "Traces only - relays to Tempo. Logs bypass it entirely (Loki's native push API instead); no metrics pipeline (Prometheus scrapes services directly).", "no auth locally"),
]
ws7 = add_sheet(
    "Infra & Credentials",
    ["Service", "Type", "Address", "Purpose", "Local credentials"],
    infra_rows,
    widths=[36, 16, 34, 55, 40],
    row_height=32,
)

for name in wb.sheetnames:
    if name != "Overview":
        wb[name].sheet_view.showGridLines = False

wb.move_sheet("Overview", offset=-len(wb.sheetnames) + 1)
wb.save(OUT)
print("wrote", OUT)

# Bake a base64 copy of the workbook into index.html so the page also works
# when opened directly (file://), where browsers block fetch("data.xlsx").
# index.html always prefers the live file when served over http; this is
# only the offline fallback.
with open(OUT, "rb") as f:
    b64 = base64.b64encode(f.read()).decode("ascii")

with open(HTML, "r", encoding="utf-8") as f:
    html = f.read()

pattern = r'const EMBEDDED_XLSX_B64 = "[^"]*";'
replacement = f'const EMBEDDED_XLSX_B64 = "{b64}";'
new_html, n = re.subn(pattern, replacement, html)
if n != 1:
    raise SystemExit(f"expected exactly one EMBEDDED_XLSX_B64 placeholder in {HTML}, found {n}")

with open(HTML, "w", encoding="utf-8") as f:
    f.write(new_html)
print(f"embedded {len(b64)} base64 chars into {HTML}")
